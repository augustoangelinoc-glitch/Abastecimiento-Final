# -*- coding: utf-8 -*-
"""
logic.py - Control de Inventarios y Abastecimiento

Reglas principales:
- La demanda mensual se construye EXCLUSIVAMENTE desde Salidas.
- Se conserva cada mes calendario del periodo de Salidas; los meses sin salida
  de un material valen 0.
- Consumo mensual promedio = consumo total / meses con consumo.
- Consumo diario = consumo mensual promedio / 30.
- Última salida y días sin movimiento salen de Salidas.
- Los materiales que aparecen en Salidas pero no están en Stock se incorporan
  a la tabla principal con stock 0.
- Lead Time: solo OC con Estado Item = COMPRADO y fechas F. Docum. / Fecha Guia
  válidas. Se calcula por material y se usa la mediana de los Lead Time válidos.
- Última compra: última OC COMPRADO por fecha F. Docum.; cantidad de esa fila.
- Stock de seguridad: Z * desviación estándar mensual / 30 * sqrt(Lead Time).
- Punto de pedido: consumo diario * Lead Time + stock de seguridad.
- Stock objetivo: consumo diario * días objetivo + stock de seguridad.
- Compra 1/2/3 meses: máximo(0, consumo mensual * meses + stock seguridad - stock).
"""

import io
import re
import unicodedata
from datetime import datetime

import numpy as np
import pandas as pd


def _normalize(x):
    if x is None:
        return ""
    s = str(x).strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]+", " ", s).strip()


def _clean_headers(df):
    df = df.copy()
    # Evita que dos columnas con el mismo nombre hagan que d[col] devuelva DataFrame.
    seen = {}
    new = []
    for c in df.columns:
        base = str(c).strip()
        n = seen.get(base, 0)
        new.append(base if n == 0 else f"{base}__{n}")
        seen[base] = n + 1
    df.columns = new
    return df


def _find_col(df, aliases, required=False):
    cols = list(df.columns)
    norm = {_normalize(c): c for c in cols}
    for a in aliases:
        if _normalize(a) in norm:
            return norm[_normalize(a)]
    # búsqueda por coincidencia contenida
    for c in cols:
        nc = _normalize(c)
        for a in aliases:
            na = _normalize(a)
            if na and (na in nc or nc in na):
                return c
    if required:
        raise ValueError(f"No se encontró una columna requerida. Alias buscados: {aliases}")
    return None


def _read_excel(file):
    try:
        file.seek(0)
    except Exception:
        pass
    df = pd.read_excel(file)
    return _clean_headers(df)


def _read_code_as_text(file, original_col):
    """Conserva códigos con ceros iniciales cuando Excel los guarda como número."""
    try:
        import openpyxl
        file.seek(0)
        wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = None
        for i, h in enumerate(headers):
            if str(h).strip() == str(original_col).strip():
                idx = i
                break
        if idx is None:
            return None
        out = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            cell = row[idx]
            value = cell.value
            if value is None:
                out.append("")
                continue
            fmt = cell.number_format or ""
            if isinstance(value, (int, np.integer)) and re.fullmatch(r"0+", fmt):
                out.append(str(value).zfill(len(fmt)))
            elif isinstance(value, float) and value.is_integer() and re.fullmatch(r"0+", fmt):
                out.append(str(int(value)).zfill(len(fmt)))
            else:
                out.append(str(value).strip())
        return out
    except Exception:
        return None


STOCK_ALIASES = {
    "codigo": ["Código", "Codigo", "Cod", "SKU", "Item", "Material", "Código material"],
    "descripcion": ["Descripción", "Descripcion", "Nombre", "Producto", "Material"],
    "unidad": ["Unidad de medida", "U.M.", "U.M", "UM", "Unidad", "U Medida"],
    "familia": ["Familia", "Categoría", "Categoria", "Grupo", "Línea", "Linea"],
    "stock": ["Stock actual", "Stock", "Existencia", "Existencias", "Saldo", "Cantidad"],
    "costo": ["Costo unitario", "Costo Unitario (S/)", "Costo", "Costo cierre mes", "Costo Cierre Mes", "V/U", "Precio"],
}

SAL_ALIASES = {
    "codigo": STOCK_ALIASES["codigo"],
    "descripcion": STOCK_ALIASES["descripcion"],
    "unidad": STOCK_ALIASES["unidad"],
    "familia": STOCK_ALIASES["familia"],
    "fecha": ["Fecha", "Fecha salida", "Fecha de salida", "Fecha movimiento", "F. Salida"],
    "cantidad": ["Cantidad salida", "Cantidad de salida", "Salida", "Consumo", "Cantidad", "Qty", "Cant salida"],
}

OC_ALIASES = {
    "codigo": STOCK_ALIASES["codigo"],
    "descripcion": STOCK_ALIASES["descripcion"],
    "fecha_doc": ["F. Docum.", "F Docum", "Fecha Documento", "Fecha Docum", "Fecha de documento"],
    "fecha_guia": ["Fecha Guia", "Fecha Guía", "Fecha de Guía", "Fecha recepción", "Fecha recepcion", "Recepción"],
    "estado_item": ["Estado Item", "Estado Ítem", "Estado item", "Estado"],
    "cantidad": ["Cantidad", "Cantidad OC", "Cant. OC", "Cantidad solicitada", "Cantidad comprada", "Qty"],
    "numero_oc": ["P.OC", "N° OC", "No OC", "Número OC", "Numero OC", "OC"],
}


def load_stock_file(file):
    df = _read_excel(file)
    mapping = {k: _find_col(df, v, k in ["codigo", "stock"]) for k, v in STOCK_ALIASES.items()}
    raw = _read_code_as_text(file, mapping["codigo"])
    out = pd.DataFrame()
    out["codigo"] = raw if raw is not None and len(raw) == len(df) else df[mapping["codigo"]].astype(str)
    out["codigo"] = out["codigo"].replace({"nan": np.nan, "None": np.nan}).astype("string").str.strip()
    out["descripcion"] = df[mapping["descripcion"]].astype("string").str.strip() if mapping["descripcion"] else ""
    out["unidad"] = df[mapping["unidad"]].astype("string").str.strip().str.upper() if mapping["unidad"] else ""
    out["familia"] = df[mapping["familia"]].astype("string").str.strip() if mapping["familia"] else ""
    out["stock_actual"] = pd.to_numeric(df[mapping["stock"]], errors="coerce").fillna(0.0)
    out["costo_unitario"] = pd.to_numeric(df[mapping["costo"]], errors="coerce").fillna(0.0) if mapping["costo"] else 0.0
    # Un solo registro por código; si hay duplicados, suma stock y conserva datos descriptivos.
    out = out.groupby("codigo", as_index=False).agg({
        "descripcion": "first", "unidad": "first", "familia": "first",
        "stock_actual": "sum", "costo_unitario": "first"
    })
    return out


def load_salidas_file(file):
    df = _read_excel(file)
    mapping = {k: _find_col(df, v, k in ["codigo", "fecha", "cantidad"]) for k, v in SAL_ALIASES.items()}
    raw = _read_code_as_text(file, mapping["codigo"])
    out = pd.DataFrame()
    out["codigo"] = raw if raw is not None and len(raw) == len(df) else df[mapping["codigo"]].astype(str)
    out["codigo"] = out["codigo"].replace({"nan": np.nan, "None": np.nan}).astype("string").str.strip()
    out["descripcion"] = df[mapping["descripcion"]].astype("string").str.strip() if mapping["descripcion"] else ""
    out["unidad"] = df[mapping["unidad"]].astype("string").str.strip().str.upper() if mapping["unidad"] else ""
    out["familia"] = df[mapping["familia"]].astype("string").str.strip() if mapping["familia"] else ""
    out["fecha"] = pd.to_datetime(df[mapping["fecha"]], errors="coerce", dayfirst=True)
    out["cantidad_salida"] = pd.to_numeric(df[mapping["cantidad"]], errors="coerce")
    out = out[out["codigo"].notna() & out["fecha"].notna() & out["cantidad_salida"].notna()]
    out = out[out["cantidad_salida"] >= 0].copy()
    return out


def load_oc_file(file):
    df = _read_excel(file)
    mapping = {k: _find_col(df, v, k in ["codigo", "fecha_doc", "fecha_guia", "estado_item"]) for k, v in OC_ALIASES.items()}
    raw = _read_code_as_text(file, mapping["codigo"])
    out = pd.DataFrame()
    out["codigo"] = raw if raw is not None and len(raw) == len(df) else df[mapping["codigo"]].astype(str)
    out["codigo"] = out["codigo"].replace({"nan": np.nan, "None": np.nan}).astype("string").str.strip()
    out["descripcion"] = df[mapping["descripcion"]].astype("string").str.strip() if mapping["descripcion"] else ""
    out["f_docum"] = pd.to_datetime(df[mapping["fecha_doc"]], errors="coerce", dayfirst=True)
    out["fecha_guia"] = pd.to_datetime(df[mapping["fecha_guia"]], errors="coerce", dayfirst=True)
    out["estado_item"] = df[mapping["estado_item"]].astype("string").str.strip().str.upper()
    out["cantidad_oc"] = pd.to_numeric(df[mapping["cantidad"]], errors="coerce") if mapping["cantidad"] else np.nan
    out["numero_oc"] = df[mapping["numero_oc"]].astype("string").str.strip() if mapping["numero_oc"] else ""
    out["lead_time"] = (out["fecha_guia"] - out["f_docum"]).dt.days
    return out[out["codigo"].notna()].copy()


def build_monthly_consumption(salidas):
    if salidas.empty:
        return pd.DataFrame(columns=["codigo", "periodo", "periodo_str", "consumo"])
    x = salidas.copy()
    x["periodo"] = x["fecha"].dt.to_period("M")
    m = x.groupby(["codigo", "periodo"], as_index=False)["cantidad_salida"].sum().rename(columns={"cantidad_salida": "consumo"})
    m["periodo_str"] = m["periodo"].astype(str)
    return m


def _classify_frequency(n, months):
    if n >= max(1, int(np.ceil(months * 0.75))):
        return "Frecuente"
    if n >= max(2, int(np.ceil(months * 0.40))):
        return "Intermitente"
    if n == 1:
        return "Ocasional"
    return "Eventual"


def _trend(values):
    y = np.asarray(values, dtype=float)
    if len(y) < 2 or np.allclose(y, y[0]):
        return "Estable"
    x = np.arange(len(y))
    slope = np.polyfit(x, y, 1)[0]
    base = max(abs(np.mean(y)), 1e-9)
    ratio = slope / base
    if ratio > 0.03:
        return "Creciente"
    if ratio < -0.03:
        return "Decreciente"
    return "Estable"


def _xyz(cv):
    if not np.isfinite(cv):
        return "Z"
    if cv <= 0.50:
        return "X"
    if cv <= 1.00:
        return "Y"
    return "Z"


def compute_analysis(stock, salidas, oc=None, z=1.65, dias_objetivo=90, fecha_corte=None):
    if oc is None:
        oc = pd.DataFrame()
    s = salidas.copy()
    if s.empty:
        raise ValueError("El archivo de Salidas no contiene registros válidos.")

    fecha_min = s["fecha"].min()
    fecha_max = fecha_corte if fecha_corte is not None else s["fecha"].max()
    fecha_max = pd.Timestamp(fecha_max)
    periodos = list(pd.period_range(fecha_min.to_period("M"), fecha_max.to_period("M"), freq="M"))
    monthly = build_monthly_consumption(s)
    codes = pd.Index(sorted(set(stock["codigo"].dropna()) | set(s["codigo"].dropna())))

    pivot = monthly.pivot_table(index="codigo", columns="periodo", values="consumo", aggfunc="sum", fill_value=0)
    pivot = pivot.reindex(index=codes, columns=periodos, fill_value=0)

    # Información maestra de Stock.
    stock_idx = stock.set_index("codigo") if not stock.empty else pd.DataFrame()
    sal_info = s.sort_values("fecha").groupby("codigo").agg(
        ultima_salida=("fecha", "max"),
        consumo_total=("cantidad_salida", "sum"),
        movimientos=("cantidad_salida", "size"),
    )

    # Lead Time por OC: solo COMPRADO + fechas válidas + 0..365 días.
    lt_map = {}
    last_buy_qty = {}
    last_buy_date = {}
    if not oc.empty:
        valid = oc[
            oc["estado_item"].fillna("").str.upper().eq("COMPRADO")
            & oc["f_docum"].notna()
            & oc["fecha_guia"].notna()
            & oc["lead_time"].between(0, 365, inclusive="both")
        ].copy()
        if not valid.empty:
            lt_stats = valid.groupby("codigo")["lead_time"].median()
            lt_map = lt_stats.to_dict()
            valid = valid.sort_values(["codigo", "f_docum", "fecha_guia"])
            last = valid.groupby("codigo", as_index=False).tail(1)
            last_buy_qty = last.set_index("codigo")["cantidad_oc"].to_dict()
            last_buy_date = last.set_index("codigo")["f_docum"].to_dict()

    rows = []
    for codigo in codes:
        serie = pivot.loc[codigo].astype(float).values if codigo in pivot.index else np.zeros(len(periodos))
        total = float(serie.sum())
        nonzero = serie[serie > 0]
        n_con = int((serie > 0).sum())
        n_total = len(serie)
        n_sin = n_total - n_con
        prom = total / n_con if n_con > 0 else 0.0
        diario = prom / 30.0
        std = float(np.std(nonzero, ddof=1)) if len(nonzero) >= 2 else 0.0
        cv = std / prom if prom > 0 else np.nan

        if codigo in stock_idx.index:
            r = stock_idx.loc[codigo]
            desc = r.get("descripcion", "")
            unidad = r.get("unidad", "")
            familia = r.get("familia", "")
            stock_actual = float(r.get("stock_actual", 0) or 0)
            costo = float(r.get("costo_unitario", 0) or 0)
        else:
            rows_sal = s[s["codigo"] == codigo]
            desc = rows_sal["descripcion"].dropna().iloc[0] if rows_sal["descripcion"].notna().any() else ""
            unidad = rows_sal["unidad"].dropna().iloc[0] if rows_sal["unidad"].notna().any() else ""
            familia = rows_sal["familia"].dropna().iloc[0] if rows_sal["familia"].notna().any() else ""
            stock_actual = 0.0
            costo = 0.0

        ultima = sal_info.loc[codigo, "ultima_salida"] if codigo in sal_info.index else pd.NaT
        dias_sin = int((fecha_max.normalize() - ultima.normalize()).days) if pd.notna(ultima) else np.nan
        cobertura = stock_actual / diario if diario > 0 else np.inf

        lt = float(lt_map.get(codigo, np.nan))
        lt_uso = lt if np.isfinite(lt) else 0.0
        ss = z * (std / 30.0) * np.sqrt(max(lt_uso, 0.0))
        pp = diario * lt_uso + ss
        stock_obj = diario * float(dias_objetivo) + ss
        compra_1 = max(0.0, diario * 30 + ss - stock_actual)
        compra_2 = max(0.0, diario * 60 + ss - stock_actual)
        compra_3 = max(0.0, diario * 90 + ss - stock_actual)

        # Rotura: stock cero + consumo histórico.
        rotura = bool(stock_actual <= 0 and total > 0)
        if total <= 0:
            situacion = "SIN CONSUMO"
        elif stock_actual <= 0:
            situacion = "ROTURA DE STOCK"
        elif stock_actual < pp:
            situacion = "BAJO PUNTO DE PEDIDO"
        elif cobertura < dias_objetivo:
            situacion = "COBERTURA INFERIOR AL OBJETIVO"
        else:
            situacion = "STOCK SUFICIENTE"

        freq = _classify_frequency(n_con, n_total)
        if rotura and freq in ("Frecuente", "Intermitente"):
            prioridad = "CRÍTICO" if freq == "Frecuente" else "ALTO"
        elif rotura:
            prioridad = "REVISAR"
        elif stock_actual < pp:
            prioridad = "ALTO"
        elif cobertura < dias_objetivo:
            prioridad = "REVISAR"
        else:
            prioridad = "NORMAL"

        if total <= 0:
            recomend = "Sin consumo histórico; no generar compra automática."
        elif rotura and freq == "Ocasional":
            recomend = f"Stock 0 con consumo ocasional ({n_con} mes con consumo). Revisar necesidad antes de comprar."
        elif rotura:
            recomend = f"ROTURA DE STOCK: {n_con} meses con consumo y stock actual 0. Abastecer según horizonte seleccionado."
        elif stock_actual < pp:
            recomend = "Stock por debajo del punto de pedido. Programar abastecimiento."
        else:
            recomend = "Stock con cobertura; revisar según evolución y punto de pedido."

        rows.append({
            "Código": str(codigo),
            "Descripción": desc,
            "Unidad de medida": unidad,
            "Familia": familia,
            "Stock actual": stock_actual,
            "Costo unitario (S/)": costo,
            "Valor del inventario (S/)": stock_actual * costo,
            "Consumo total": total,
            "Valor de salidas (S/)": total * costo,
            "Última salida": ultima,
            **{str(p): float(serie[i]) for i, p in enumerate(periodos)},
            "Meses con consumo": n_con,
            "Meses sin consumo": n_sin,
            "Consumo mensual promedio": prom,
            "Consumo diario": diario,
            "Días sin movimiento": dias_sin,
            "Cobertura (días)": cobertura,
            "Lead Time (días)": lt,
            "Tipo de consumo": freq,
            "Situación de stock": situacion,
            "Rotura de stock": "VERDADERO" if rotura else "FALSO",
            "Tendencia del consumo": _trend(serie),
            "Coeficiente de variación": cv,
            "Nivel de variabilidad": "Baja" if np.isfinite(cv) and cv <= .50 else ("Media" if np.isfinite(cv) and cv <= 1 else "Alta"),
            "Anomalía de consumo": "No",
            "ABC acumulado (%)": np.nan,
            "Clasificación ABC": "",
            "Clasificación XYZ": _xyz(cv),
            "Stock de seguridad": ss,
            "Punto de pedido": pp,
            "Stock objetivo": stock_obj,
            "Cantidad a abastecer": compra_1,
            "Última compra": float(last_buy_qty.get(codigo, np.nan)) if pd.notna(last_buy_qty.get(codigo, np.nan)) else np.nan,
            "Fecha última compra": last_buy_date.get(codigo, pd.NaT),
            "Días desde última compra": ((fecha_max.normalize() - pd.Timestamp(last_buy_date[codigo]).normalize()).days if codigo in last_buy_date else np.nan),
            "Cobertura última compra (meses)": (float(last_buy_qty[codigo]) / prom if codigo in last_buy_qty and prom > 0 else np.nan),
            "Compra para 1 mes": compra_1,
            "Duración después de comprar 1 mes (días)": (stock_actual + compra_1) / diario if diario > 0 else np.inf,
            "Compra para 2 meses": compra_2,
            "Duración después de comprar 2 meses (días)": (stock_actual + compra_2) / diario if diario > 0 else np.inf,
            "Compra para 3 meses": compra_3,
            "Duración después de comprar 3 meses (días)": (stock_actual + compra_3) / diario if diario > 0 else np.inf,
            "Cuándo comprar": "AHORA" if (stock_actual <= pp and total > 0) else "ESPERAR",
            "Metodología utilizada": "Promedio de meses con consumo + variabilidad + Lead Time de OC COMPRADO",
            "Diagnóstico": f"{situacion}. {freq}. {n_con} meses con consumo de {n_total}.",
            "Prioridad": prioridad,
            "Recomendación": recomend,
        })

    result = pd.DataFrame(rows)

    # ABC por valor de salidas. Si todos los valores son 0, no inventar diferencias:
    # se deja 100% acumulado y C para mantener el comportamiento de la versión previa.
    vals = result["Valor de salidas (S/)"].fillna(0).clip(lower=0)
    if vals.sum() > 0:
        order = vals.sort_values(ascending=False).index
        cum = vals.loc[order].cumsum() / vals.sum() * 100
        result.loc[order, "ABC acumulado (%)"] = cum.values
        result.loc[order, "Clasificación ABC"] = np.where(cum.values <= 80, "A", np.where(cum.values <= 95, "B", "C"))
    else:
        result["ABC acumulado (%)"] = 100.0
        result["Clasificación ABC"] = "C"

    # Redondeos solo de presentación; los cálculos conservan precisión internamente.
    result = result.sort_values(["Prioridad", "Consumo mensual promedio"], ascending=[True, False], na_position="last").reset_index(drop=True)
    return result, monthly, periodos


def build_export_excel(result, monthly, issues, params, path):
    result = result.copy()
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        result.to_excel(writer, sheet_name="Principal", index=False)
        if monthly.empty:
            pd.DataFrame({"Aviso": ["No hay consumo mensual"]}).to_excel(writer, sheet_name="Consumo Mensual", index=False)
        else:
            p = monthly.pivot_table(index="codigo", columns="periodo_str", values="consumo", aggfunc="sum", fill_value=0).reset_index()
            p.rename(columns={"codigo": "Código"}, inplace=True)
            p.to_excel(writer, sheet_name="Consumo Mensual", index=False)

        rot = result[result["Rotura de stock"] == "VERDADERO"].copy()
        rot.to_excel(writer, sheet_name="Roturas de Stock", index=False)

        compras = result[result["Compra para 1 mes"] > 0].copy()
        compras.to_excel(writer, sheet_name="Recomendación de Compra", index=False)

        oc_info = pd.DataFrame()
        if params.get("oc") is not None and not params["oc"].empty:
            oc_info = params["oc"].copy()
        if oc_info.empty:
            pd.DataFrame({"Aviso": ["No se cargó OC o no hubo registros válidos."]}).to_excel(writer, sheet_name="Ordenes de Compra", index=False)
        else:
            oc_info.to_excel(writer, sheet_name="Ordenes de Compra", index=False)

        if issues:
            start = 0
            rows = []
            for k, v in issues.items():
                if k.startswith("_"):
                    continue
                if isinstance(v, pd.DataFrame) and not v.empty:
                    rows.append(pd.DataFrame({"Tipo": [k], "Cantidad": [len(v)]}))
            pd.concat(rows, ignore_index=True).to_excel(writer, sheet_name="Validación", index=False) if rows else pd.DataFrame({"Resultado": ["Sin inconsistencias"]}).to_excel(writer, sheet_name="Validación", index=False)


def validate_data(stock, salidas, oc=None):
    issues = {}
    cs = set(stock["codigo"].dropna())
    cl = set(salidas["codigo"].dropna())
    issues["Stock sin salidas"] = stock[stock["codigo"].isin(cs - cl)][["codigo", "descripcion", "familia", "unidad"]]
    issues["Salidas sin stock"] = salidas[salidas["codigo"].isin(cl - cs)][["codigo", "descripcion", "familia", "unidad"]].drop_duplicates()
    issues["_total"] = sum(len(v) for v in issues.values() if isinstance(v, pd.DataFrame))
    return issues
